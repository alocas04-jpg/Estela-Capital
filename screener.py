"""
Screener Moats — Estela Capital
Detecta empresas de calidad con derating significativo y evalúa si el moat sigue intacto.

Uso:
    python screener.py                    # corre todo el universo
    python screener.py --test 10          # prueba con las primeras 10 empresas
    python screener.py --threshold 30     # umbral de derating personalizado (default: 40%)
"""

import json
import os
import re
import sys
import time
import argparse
import subprocess
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from datetime import date
from pathlib import Path

# ---------------------------------------------------------------------------
# Rutas
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent
CLAUDE_DIR = BASE_DIR.parents[1]   # .claude/
ROOT_DIR   = BASE_DIR.parents[2]   # 1 - Analisis y Gestión/
# Busca companies.json primero en la carpeta del script (para GitHub Actions), luego en la ruta original
_local_companies = BASE_DIR / "companies.json"
_remote_companies = CLAUDE_DIR / "AI DIGEST" / "backend" / "config" / "companies.json"
COMPANIES_JSON = _local_companies if _local_companies.exists() else _remote_companies
# Busca la API key en este orden: tools/.env → raíz/.env → variable de entorno del sistema
ENV_FILES = [CLAUDE_DIR / "tools" / ".env", ROOT_DIR / ".env"]
OUTPUT_DIR = BASE_DIR / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

# Archivos de tracking
EV_HISTORY_FILE = BASE_DIR / "ev_history.json"
TRACKING_FILE   = BASE_DIR / "tracking.json"

# ---------------------------------------------------------------------------
# Cargar .env manualmente (sin python-dotenv)
# ---------------------------------------------------------------------------
def load_env(env_path: Path):
    if not env_path.exists():
        return
    with open(env_path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, val = line.partition("=")
            os.environ.setdefault(key.strip(), val.strip())

for _env in ENV_FILES:
    load_env(_env)

# ---------------------------------------------------------------------------
# Imports de terceros (después de confirmar instalación)
# ---------------------------------------------------------------------------
try:
    import yfinance as yf
    import pandas as pd
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    try:
        import anthropic
    except ModuleNotFoundError:
        anthropic = None
except ModuleNotFoundError as e:
    print(f"[ERROR] Falta una dependencia: {e}")
    print("Ejecuta: pip install yfinance openpyxl pandas")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Gemini client wrapper — reemplaza Anthropic cuando no hay créditos
# ---------------------------------------------------------------------------
import urllib.request as _urllib_req
import urllib.error as _urllib_err

class GroqClient:
    """Wrapper que imita anthropic.Anthropic usando Groq (gratis, 14400 req/día)."""
    MODEL = "llama-3.1-8b-instant"

    def __init__(self, api_key: str):
        self.api_key = api_key
        self.messages = self

    def create(self, model=None, max_tokens=900, system="", messages=None, **kwargs):
        import time as _time
        msgs = []
        if system:
            msgs.append({"role": "system", "content": system})
        if messages:
            msgs.extend(messages)
        url = "https://api.groq.com/openai/v1/chat/completions"
        body = json.dumps({
            "model": self.MODEL,
            "messages": msgs,
            "max_tokens": max_tokens,
            "temperature": 0.2,
        }).encode("utf-8")
        req = _urllib_req.Request(url, data=body, headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {self.api_key}",
            "User-Agent": "estela-screener/1.0",
        })
        for _attempt in range(3):
            try:
                with _urllib_req.urlopen(req, timeout=30) as resp:
                    data = json.loads(resp.read().decode("utf-8"))
                text = data["choices"][0]["message"]["content"].strip()
                # Strip markdown code blocks if model wraps JSON
                if text.startswith("```"):
                    lines = text.split("\n")
                    inner = "\n".join(lines[1:-1]) if lines[-1].strip() == "```" else "\n".join(lines[1:])
                    text = inner.strip()
                class _Resp:
                    content = [type("_P", (), {"text": text})()]
                return _Resp()
            except Exception as e:
                _msg = str(e)
                if "429" in _msg or "rate" in _msg.lower():
                    _wait = 30 * (2 ** _attempt)  # 30s, 60s, 120s
                    print(f"  [GROQ] Rate limit, esperando {_wait}s...")
                    _time.sleep(_wait)
                else:
                    print(f"  [GROQ ERROR] {type(e).__name__}: {_msg[:200]}")
                    break
        class _Err:
            content = [type("_P", (), {"text": "{}"})()]
        return _Err()

# ---------------------------------------------------------------------------
# Configuración
# ---------------------------------------------------------------------------
DERATING_THRESHOLD = 40          # % mínimo de caída para pasar al análisis de moat
CLAUDE_MODEL = "claude-sonnet-4-6"
BATCH_DELAY = 7.0                # segundos entre llamadas AI (Groq 8b: 20000 TPM, ~2000 tokens/call = max 10/min)
YF_PERIOD = "5y"                 # histórico para calcular P/E medio

# Mapeo exchange → sufijo Yahoo Finance
EXCHANGE_SUFFIX = {
    "XLON": ".L",
    "XETR": ".DE",
    "XPAR": ".PA",
    "XAMS": ".AS",
    "XSWX": ".SW",
    "XMIL": ".MI",
    "XMAD": ".MC",
    "XNAS": "",        # NASDAQ — sin sufijo
    "XNYS": "",        # NYSE — sin sufijo
    "XTKS": ".T",
    "XHKG": ".HK",
    "XASX": ".AX",
    "XBOM": ".BO",
    "XNSE": ".NS",
    "BVMF": ".SA",
    "XKRX": ".KS",
    "XTAI": ".TW",
    "XSHG": ".SS",
    "XSHE": ".SZ",
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def resolve_ticker(name: str, ticker: str, exchange: str) -> str:
    """Construye el ticker Yahoo Finance a partir del ticker base y el exchange."""
    suffix = EXCHANGE_SUFFIX.get(exchange, "")
    return f"{ticker}{suffix}"


# ---------------------------------------------------------------------------
# IMPROVEMENT 2 — EV/EBITDA history helpers
# ---------------------------------------------------------------------------
def load_ev_history() -> dict:
    if EV_HISTORY_FILE.exists():
        try:
            with open(EV_HISTORY_FILE, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def save_ev_history(history: dict):
    with open(EV_HISTORY_FILE, "w", encoding="utf-8") as f:
        json.dump(history, f, ensure_ascii=False, indent=2)


# ---------------------------------------------------------------------------
# IMPROVEMENT 3 — Tracking history helpers
# ---------------------------------------------------------------------------
def load_tracking() -> dict:
    if TRACKING_FILE.exists():
        try:
            with open(TRACKING_FILE, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def save_tracking(tracking: dict):
    with open(TRACKING_FILE, "w", encoding="utf-8") as f:
        json.dump(tracking, f, ensure_ascii=False, indent=2)


def update_tracking(ticker: str, drop_pct: float, vale_investigar: bool, conviction: int, tracking: dict) -> str:
    """
    Updates tracking.json entry for ticker.
    Returns tendencia: 'Nueva' | 'Profundizando' | 'Recuperando' | 'Estable'
    """
    today_str = date.today().strftime("%Y-%m-%d")
    entry = {
        "date": today_str,
        "drop_pct": drop_pct,
        "vale_investigar": vale_investigar,
        "conviction": conviction,
    }

    if ticker not in tracking:
        tracking[ticker] = {
            "first_seen": today_str,
            "entries": [entry],
        }
        return "Nueva"

    entries = tracking[ticker]["entries"]
    tendencia = "Estable"
    if entries:
        last_drop = entries[-1].get("drop_pct", drop_pct)
        diff = drop_pct - last_drop
        if diff > 2:
            tendencia = "Profundizando"
        elif diff < -2:
            tendencia = "Recuperando"
        else:
            tendencia = "Estable"

    # Avoid duplicate entry for same date
    if not entries or entries[-1].get("date") != today_str:
        entries.append(entry)

    return tendencia


# ---------------------------------------------------------------------------
# IMPROVEMENT 4 — Perplexity news context
# ---------------------------------------------------------------------------
def get_news_context(company_name: str, perplexity_key: str) -> str:
    """Returns 3-5 bullet points of recent news (last 2 weeks) about the company."""
    try:
        import requests
        headers = {
            "Authorization": f"Bearer {perplexity_key}",
            "Content-Type": "application/json",
        }
        payload = {
            "model": "sonar",
            "messages": [
                {
                    "role": "user",
                    "content": (
                        f"What are the main news and events affecting {company_name} stock "
                        f"in the last 2 weeks? Give me 3-5 bullet points, factual, focused on "
                        f"what is causing the stock decline. Be concise."
                    ),
                }
            ],
            "search_recency_filter": "week",
        }
        resp = requests.post(
            "https://api.perplexity.ai/chat/completions",
            headers=headers,
            json=payload,
            timeout=30,
        )
        resp.raise_for_status()
        return resp.json()["choices"][0]["message"]["content"]
    except Exception as e:
        print(f"  [WARN] Perplexity news para {company_name}: {e}")
        return ""


# ---------------------------------------------------------------------------
# IMPROVEMENT 5 — Vault integration helper
# ---------------------------------------------------------------------------
def get_vault_excerpt(company_name: str) -> tuple:
    """
    Returns (found: bool, excerpt: str) — first 500 chars of any .md file
    in CLAUDE_DIR / "Estela's Vault" / "Companies" / company_name.
    """
    vault_dir = CLAUDE_DIR / "Estela's Vault" / "Companies" / company_name
    if not vault_dir.exists():
        return False, ""
    md_files = list(vault_dir.glob("*.md"))
    if not md_files:
        return False, ""
    try:
        with open(md_files[0], encoding="utf-8") as f:
            content = f.read(500)
        return True, content
    except Exception:
        return False, ""


# ---------------------------------------------------------------------------
# get_derating_signals — with IMPROVEMENT 2 (EV/EBITDA)
# ---------------------------------------------------------------------------
def get_derating_signals(yf_ticker: str, ev_history: dict = None) -> dict:
    """
    Obtiene señales de derating vía yfinance:
    - Caída desde máximo 52 semanas
    - Compresión de P/E vs. media 5 años
    - EV/EBITDA actual + histórico (mejora 2)
    Devuelve dict con los campos o None si falla.
    """
    if ev_history is None:
        ev_history = {}

    try:
        stock = yf.Ticker(yf_ticker)
        info = stock.info

        # Precio actual y máximo 52 semanas
        current_price = info.get("currentPrice") or info.get("regularMarketPrice")
        high_52w = info.get("fiftyTwoWeekHigh")

        if not current_price or not high_52w or high_52w == 0:
            return {"error": "precio no disponible"}

        drop_pct = round((high_52w - current_price) / high_52w * 100, 1)

        # P/E actual
        pe_current = info.get("trailingPE") or info.get("forwardPE")

        # P/E medio histórico
        pe_mean_5y = None
        pe_compression = None
        try:
            hist = stock.history(period=YF_PERIOD, interval="3mo")
            earnings = stock.earnings_history
        except Exception:
            pass

        if pe_current and pe_mean_5y:
            pe_compression = round((pe_mean_5y - pe_current) / pe_mean_5y * 100, 1)
        else:
            pe_compression = None

        # --- EV/EBITDA (Improvement 2) ---
        ev_ebitda_current = info.get("enterpriseToEbitda")
        if ev_ebitda_current is not None:
            ev_ebitda_current = round(float(ev_ebitda_current), 2)

        # Append to history
        ticker_key = yf_ticker.upper()
        if ev_ebitda_current is not None:
            if ticker_key not in ev_history:
                ev_history[ticker_key] = []
            ev_history[ticker_key].append(ev_ebitda_current)

        # Calculate mean and discount if >=4 data points
        ev_ebitda_5y_mean = None
        ev_ebitda_discount_pct = None
        history_vals = ev_history.get(ticker_key, [])
        if len(history_vals) >= 4:
            ev_ebitda_5y_mean = round(sum(history_vals) / len(history_vals), 2)
            if ev_ebitda_current and ev_ebitda_5y_mean and ev_ebitda_5y_mean != 0:
                ev_ebitda_discount_pct = round(
                    (ev_ebitda_5y_mean - ev_ebitda_current) / ev_ebitda_5y_mean * 100, 1
                )

        # Price 1 month ago (≈22 trading days) for performance indicator
        price_1m_ago = None
        try:
            hist = stock.history(period="2mo", interval="1d")
            if len(hist) >= 22:
                price_1m_ago = round(float(hist["Close"].iloc[-22]), 2)
        except Exception:
            pass

        # Quality & valuation metrics from info
        roe          = info.get("returnOnEquity")          # e.g. 0.25 = 25%
        gross_margin = info.get("grossMargins")            # e.g. 0.60 = 60%
        revenue_growth = info.get("revenueGrowth")        # 1y YoY e.g. 0.12 = 12%
        fcf          = info.get("freeCashflow")
        mkt_cap      = info.get("marketCap")
        fcf_yield    = round(fcf / mkt_cap * 100, 1) if fcf and mkt_cap and mkt_cap > 0 else None
        dividend_yield = info.get("dividendYield")        # e.g. 0.02 = 2%
        beta         = info.get("beta")

        # Debt / EBITDA
        net_debt     = None
        debt_ebitda  = None
        try:
            total_debt = info.get("totalDebt") or 0
            cash       = info.get("totalCash") or 0
            ebitda     = info.get("ebitda")
            net_debt   = total_debt - cash
            if ebitda and ebitda != 0:
                debt_ebitda = round(net_debt / ebitda, 1)
        except Exception:
            pass

        # Analyst consensus
        analyst_target  = info.get("targetMeanPrice")
        analyst_upside  = None
        if analyst_target and current_price and current_price > 0:
            analyst_upside = round((analyst_target - current_price) / current_price * 100, 1)
        analyst_buy     = info.get("numberOfAnalystOpinions")
        rec_key         = info.get("recommendationKey", "")   # "buy","hold","sell","strong_buy"

        # Business description — truncated to 300 chars for email
        raw_desc = info.get("longBusinessSummary", "") or ""
        description = (raw_desc[:297] + "...") if len(raw_desc) > 300 else raw_desc

        # Price history for sparkline chart (12 months of daily closes)
        price_history = None
        try:
            hist_1y = stock.history(period="1y", interval="1d")
            if len(hist_1y) > 10:
                price_history = [round(float(p), 2) for p in hist_1y["Close"].tolist()]
        except Exception:
            pass

        return {
            "current_price": round(current_price, 2),
            "price_1m_ago": price_1m_ago,
            "high_52w": round(high_52w, 2),
            "drop_from_high_pct": drop_pct,
            "pe_current": round(pe_current, 1) if pe_current else None,
            "pe_mean_5y": pe_mean_5y,
            "pe_compression_pct": pe_compression,
            "currency": info.get("currency", ""),
            "market_cap": mkt_cap,
            "ev_ebitda_current": ev_ebitda_current,
            "ev_ebitda_5y_mean": ev_ebitda_5y_mean,
            "ev_ebitda_discount_pct": ev_ebitda_discount_pct,
            "roe": round(roe * 100, 1) if roe else None,
            "gross_margin": round(gross_margin * 100, 1) if gross_margin else None,
            "revenue_growth_1y": round(revenue_growth * 100, 1) if revenue_growth else None,
            "fcf_yield": fcf_yield,
            "dividend_yield": round(dividend_yield * 100, 2) if dividend_yield else None,
            "beta": round(beta, 2) if beta else None,
            "net_debt_ebitda": debt_ebitda,
            "analyst_target": round(analyst_target, 2) if analyst_target else None,
            "analyst_upside_pct": analyst_upside,
            "analyst_count": analyst_buy,
            "recommendation": rec_key,
            "description": description,
            "sector": info.get("sector", "") or info.get("industry", ""),
            "price_history": price_history,
        }
    except Exception as ex:
        return {"error": str(ex)}


def passes_derating_threshold(signals: dict, threshold: float) -> bool:
    """Devuelve True si la empresa supera el umbral de derating."""
    if "error" in signals:
        return False
    drop = signals.get("drop_from_high_pct", 0) or 0
    compression = signals.get("pe_compression_pct", 0) or 0
    return drop >= threshold or compression >= threshold


# ---------------------------------------------------------------------------
# Análisis de moat con Claude — IMPROVEMENT 1 (conviction score)
# ---------------------------------------------------------------------------
FEW_SHOT_EXAMPLES = """
// Ejemplo 1 — Adidas (oportunidad clara)
{
  "moat_type": "intangibles",
  "moat_strength": "wide",
  "moat_intact": true,
  "narrativa_negativa": "Ruptura contrato Yeezy + acumulación inventario + incertidumbre exposición China consumidor",
  "reversibilidad": "gestion",
  "vale_investigar": true,
  "razon": "El moat de marca global con pricing power no se ha erosionado — el mercado castiga un error de gestión puntual y un ciclo de inventario, no un deterioro estructural",
  "comparable": "Nike 2023",
  "conviction_score": 8,
  "conviction_razon": "Marca icónica con pricing power demostrado, el problema Yeezy es resoluble y el ciclo de inventario es temporal."
}

// Ejemplo 2 — Edenred (oportunidad clara)
{
  "moat_type": "switching_costs",
  "moat_strength": "wide",
  "moat_intact": true,
  "narrativa_negativa": "Regulación francesa limitando float de tickets restaurante + presión política en márgenes",
  "reversibilidad": "ciclo",
  "vale_investigar": true,
  "razon": "Switching costs estructurales en todos los mercados excepto Francia — el mercado sobrepondera un riesgo regulatorio local y olvida la fortaleza del modelo en el resto del mundo",
  "comparable": "Pluxee",
  "conviction_score": 7,
  "conviction_razon": "Modelo de negocio con switching costs muy altos y crecimiento internacional sólido, el riesgo francés está sobredescontado."
}

// Ejemplo 3 — Costar (descartar)
{
  "moat_type": "network_effects",
  "moat_strength": "narrow",
  "moat_intact": false,
  "narrativa_negativa": "Disrupción IA en búsqueda inmobiliaria comercial + reducción visibilidad post-pandemia en demanda CRE",
  "reversibilidad": "estructural",
  "vale_investigar": false,
  "razon": "El riesgo de disrupción no es temporal — la IA puede erosionar el moat de datos y red en CRE con visibilidad insuficiente sobre el outcome. Sin margen de seguridad claro.",
  "comparable": null,
  "conviction_score": 2,
  "conviction_razon": "Riesgo estructural de disrupción por IA sin respuesta clara, el moat de datos es vulnerable."
}
"""

SYSTEM_PROMPT = """Eres un analista de inversión value especializado en moats cualitativos según el framework de Pat Dorsey (The Little Book That Builds Wealth). Tu trabajo es evaluar si una empresa tiene una ventaja competitiva duradera y si esa ventaja sigue intacta a pesar del derating que está sufriendo en bolsa.

Los 5 tipos de moat de Dorsey:
1. Intangible assets — marcas con pricing power real, patentes esenciales, licencias regulatorias
2. Switching costs — coste alto de cambiar de proveedor (ERP, datos históricos, integraciones críticas)
3. Network effects — producto más valioso cuantos más usuarios (plataformas, marketplaces, estándares)
4. Cost advantage — escala, proceso o localización estructuralmente diferencial (no eficiencia temporal)
5. Efficient scale — mercado nicho donde un segundo entrante destruiría la economía para todos

Filosofía de Estela Capital: invertimos en empresas de calidad cuando el mercado duda de ellas. Nos interesan derátings causados por narrativas temporales (ciclo, error de gestión puntual, regulación sobredescontada) — NO nos interesan deterioros estructurales del moat.

Responde ÚNICAMENTE en JSON válido, sin texto adicional, sin markdown, sin bloques de código."""


def analyze_moat(client: anthropic.Anthropic, company: dict, signals: dict) -> dict:
    """Llama a Claude para analizar el moat. Devuelve el dict JSON o un dict de error."""
    drop = signals.get("drop_from_high_pct", "N/A")
    pe_compression = signals.get("pe_compression_pct", "N/A")
    pe_compression_str = f"{pe_compression}%" if pe_compression is not None else "N/D"

    user_prompt = f"""A continuación tienes tres ejemplos de cómo responder:

{FEW_SHOT_EXAMPLES}

Ahora analiza esta empresa:

Empresa: {company['name']}
Sector: {company.get('sector', 'N/D')}
Geografía: {company.get('geo', 'N/D')}
Caída desde máximos 52 semanas: {drop}%
Compresión de múltiplo P/E vs. histórico: {pe_compression_str}

Devuelve exactamente este JSON:
{{
  "moat_type": "intangibles | switching_costs | network_effects | cost_advantage | efficient_scale | none",
  "moat_strength": "wide | narrow | none",
  "moat_intact": true | false,
  "narrativa_negativa": "qué teme el mercado exactamente, en 1-2 frases",
  "reversibilidad": "ciclo | gestion | estructural | mixto",
  "vale_investigar": true | false,
  "razon": "una frase que explique el veredicto",
  "comparable": "empresa similar que hayamos analizado antes si aplica, si no: null",
  "conviction_score": 7,
  "conviction_razon": "una frase que explique la puntuación de convicción (1=mínima, 10=máxima oportunidad Estela)",
  "catalysts": ["catalizador 1 en 1 frase", "catalizador 2 en 1 frase"],
  "risks": ["riesgo principal en 1 frase", "riesgo secundario en 1 frase"],
  "entry_price_comment": "rango de precio atractivo de entrada o señal que confirmaría la tesis, en 1 frase"
}}"""

    try:
        response = client.messages.create(
            model=CLAUDE_MODEL,
            max_tokens=900,
            system=SYSTEM_PROMPT,
            messages=[{"role": "user", "content": user_prompt}],
        )
        raw = response.content[0].text.strip()
        # Limpiar posibles bloques markdown si Claude los incluye
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        return json.loads(raw)
    except json.JSONDecodeError as e:
        return {"error": f"JSON inválido: {e}", "raw": raw[:200] if 'raw' in dir() else ""}
    except Exception as e:
        return {"error": str(e)}


# ---------------------------------------------------------------------------
# IMPROVEMENT 6 — Active universe expansion via Perplexity
# ---------------------------------------------------------------------------
def suggest_new_companies_active(client: anthropic.Anthropic, perplexity_key: str, existing_names: list) -> list:
    """Uses Perplexity to search for quality compounders, then Claude to select/format the best 20."""
    existing_sample = ", ".join(existing_names[:30])

    # Step 1: Perplexity search for quality compounders
    perplexity_findings = ""
    try:
        import requests
        headers = {
            "Authorization": f"Bearer {perplexity_key}",
            "Content-Type": "application/json",
        }
        payload = {
            "model": "sonar",
            "messages": [
                {
                    "role": "user",
                    "content": (
                        f"List quality compounder companies with strong moats from MSCI World Quality index, "
                        f"S&P500, Stoxx600 — companies with high ROCE (>15%), recurring revenues, strong brands "
                        f"or switching costs, not in this list: {existing_sample}. "
                        f"Give me 30-40 specific company names with tickers and a brief reason for each. "
                        f"Focus on: hidden champions, niche leaders, B2B software, luxury goods, healthcare devices, "
                        f"specialty chemicals, professional services with high switching costs."
                    ),
                }
            ],
            "search_recency_filter": "month",
        }
        resp = requests.post(
            "https://api.perplexity.ai/chat/completions",
            headers=headers,
            json=payload,
            timeout=30,
        )
        resp.raise_for_status()
        perplexity_findings = resp.json()["choices"][0]["message"]["content"]
        print(f"  [Perplexity] Obtenidos datos de mercado para sugerencias activas")
    except Exception as e:
        print(f"  [WARN] Perplexity para sugerencias activas falló: {e}")
        # Fallback to Claude-only suggestions
        return suggest_new_companies(client, existing_names)

    # Step 2: Claude selects best 20 from Perplexity findings
    prompt = f"""Eres analista senior de Estela Capital, un fondo de inversión value europeo con filosofía de largo plazo.

Filosofía de inversión de Estela Capital:
- Buscamos quality compounders globales: negocios que pueden reinvertir capital a altas tasas de retorno durante muchos años
- Moat duradero según Dorsey: marcas con pricing power real, switching costs estructurales, network effects, cost advantage o efficient scale
- Preferimos negocios con ingresos recurrentes, baja intensidad de capital, márgenes altos y estables
- Nos gustan líderes de nicho globales (hidden champions europeos, líderes asiáticos de consumo, plataformas B2B con switching costs)
- Evitamos: negocios cíclicos sin moat, commodities, bancos, utilities, empresas con deuda excesiva, negocios disruptibles por IA sin respuesta clara

El universo actual ya incluye empresas como: {existing_sample}...

Un analista ha encontrado estas empresas candidatas con datos de mercado recientes:

{perplexity_findings}

Selecciona las 20 mejores que NO estén ya en el universo y que encajen perfectamente con la filosofía de Estela. Incluye:
- 6 europeas (hidden champions, líderes de nicho)
- 5 US (quality compounders con moat claro)
- 5 asiáticas (consumo, healthcare, tecnología con moat)
- 4 de cualquier geografía que sean ideas especialmente interesantes

Devuelve ÚNICAMENTE un JSON array válido, sin texto adicional, sin markdown:
[
  {{
    "nombre": "nombre completo de la empresa",
    "ticker_yahoo": "ticker en formato Yahoo Finance (ej: RWS.L, LOTUS.BR, ROK)",
    "geografia": "Europa | US | Asia | Latam",
    "sector": "sector en inglés",
    "moat_type": "intangibles | switching_costs | network_effects | cost_advantage | efficient_scale",
    "moat_strength": "wide | narrow",
    "por_que_añadir": "una frase explicando por qué encaja con la filosofía de Estela Capital"
  }}
]"""

    try:
        response = client.messages.create(
            model=CLAUDE_MODEL,
            max_tokens=4096,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = response.content[0].text.strip()
        if "```" in raw:
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        start = raw.find("[")
        end   = raw.rfind("]") + 1
        if start != -1 and end > start:
            raw = raw[start:end]
        raw = re.sub(r'//[^\n]*', '', raw)
        raw = re.sub(r',\s*([\]}])', r'\1', raw)
        return json.loads(raw)
    except Exception as e:
        print(f"  [WARN] No se pudieron generar sugerencias activas: {e}")
        return []


def suggest_new_companies(client: anthropic.Anthropic, existing_names: list) -> list:
    """Pide a Claude 20 sugerencias de nuevas empresas para el universo (fallback sin Perplexity)."""
    existing_sample = ", ".join(existing_names[:40])
    prompt = f"""Eres analista senior de Estela Capital, un fondo de inversión value europeo con filosofía de largo plazo.

Filosofía de inversión de Estela Capital:
- Buscamos quality compounders globales: negocios que pueden reinvertir capital a altas tasas de retorno durante muchos años
- Moat duradero según Dorsey: marcas con pricing power real, switching costs estructurales, network effects, cost advantage o efficient scale
- Preferimos negocios con ingresos recurrentes, baja intensidad de capital, márgenes altos y estables
- Nos gustan líderes de nicho globales (hidden champions europeos, líderes asiáticos de consumo, plataformas B2B con switching costs)
- Ejemplos del tipo de empresa que nos encanta: EssilorLuxottica, Hermès, Wolters Kluwer, LVMH, Moncler, Rightmove, Diploma PLC, Lotus Bakeries, Rational AG, Straumann, Temenos, Dassault Systèmes, Roper Technologies, Veeva Systems, Tyler Technologies
- Evitamos: negocios cíclicos sin moat, commodities, bancos, utilities, empresas con deuda excesiva, negocios disruptibles por IA sin respuesta clara

El universo actual ya incluye empresas como: {existing_sample}...

Propón 20 empresas nuevas que NO estén en esa lista y que encajen perfectamente con la filosofía de Estela. Incluye una mezcla de:
- 6 empresas europeas (hidden champions, líderes de nicho)
- 5 empresas US (quality compounders con moat claro)
- 5 empresas asiáticas (consumo, healthcare, tecnología con moat)
- 4 empresas de cualquier geografía que sean ideas especialmente interesantes

Devuelve ÚNICAMENTE un JSON array válido, sin texto adicional, sin markdown:
[
  {{
    "nombre": "nombre completo de la empresa",
    "ticker_yahoo": "ticker en formato Yahoo Finance (ej: RWS.L, LOTUS.BR, ROK)",
    "geografia": "Europa | US | Asia | Latam",
    "sector": "sector en inglés",
    "moat_type": "intangibles | switching_costs | network_effects | cost_advantage | efficient_scale",
    "moat_strength": "wide | narrow",
    "por_que_añadir": "una frase explicando por qué encaja con la filosofía de Estela Capital"
  }}
]"""

    try:
        response = client.messages.create(
            model=CLAUDE_MODEL,
            max_tokens=4096,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = response.content[0].text.strip()
        # Strip markdown code fences
        if "```" in raw:
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        # Extract just the JSON array
        start = raw.find("[")
        end   = raw.rfind("]") + 1
        if start != -1 and end > start:
            raw = raw[start:end]
        # Strip JS-style comments (// ...) that Claude sometimes adds
        raw = re.sub(r'//[^\n]*', '', raw)
        # Strip trailing commas before ] or }
        raw = re.sub(r',\s*([\]}])', r'\1', raw)
        return json.loads(raw)
    except Exception as e:
        print(f"  [WARN] No se pudieron generar sugerencias: {e}")
        return []


# ---------------------------------------------------------------------------
# Ranking de candidatas
# ---------------------------------------------------------------------------
def rank_candidates(client: anthropic.Anthropic, candidates: list) -> dict:
    """
    Toma todas las empresas con vale_investigar=True y devuelve un dict
    {nombre: {"prioridad": 1|2|3, "destacada": bool, "razon_prioridad": str}}
    Prioridad 1 = más interesante para Estela.
    """
    if not candidates:
        return {}

    lista = ""
    for i, r in enumerate(candidates, 1):
        m = r.get("moat", {})
        s = r.get("signals", {})
        lista += (f"{i}. {r['name']} | {r['geo']} | {r['sector']} | "
                  f"moat: {m.get('moat_type','')} ({m.get('moat_strength','')}) | "
                  f"caida: {s.get('drop_from_high_pct','')}% | "
                  f"reversibilidad: {m.get('reversibilidad','')} | "
                  f"conviction: {m.get('conviction_score','')} | "
                  f"{m.get('razon','')}\n")

    prompt = f"""Eres analista senior de Estela Capital, fondo de inversión value europeo.

Filosofía: quality compounders globales con moat duradero. Preferimos:
- Reversibilidad de gestión o ciclo sobre mixta o estructural
- Moat wide sobre narrow
- Negocios con ingresos recurrentes, márgenes altos, poca intensidad de capital
- Líderes de nicho globales, switching costs estructurales, marcas con pricing power real
- Caídas grandes con narrativa claramente temporal son las mejores oportunidades

Estas son las empresas que han pasado el filtro de derating y tienen moat intacto esta semana:

{lista}

Rankéalas de más a menos interesante para Estela Capital. Marca como "destacada" las top 3 que creas que encajan mejor con nuestra filosofía y tienen mayor potencial de ser una gran idea de inversión.

Devuelve ÚNICAMENTE un JSON array ordenado de mayor a menor prioridad, sin texto adicional:
[
  {{
    "nombre": "nombre exacto como aparece arriba",
    "prioridad": 1,
    "destacada": true,
    "razon_prioridad": "una frase corta explicando por qué es la más interesante para Estela"
  }}
]"""

    try:
        response = client.messages.create(
            model=CLAUDE_MODEL,
            max_tokens=2048,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = response.content[0].text.strip()
        if "```" in raw:
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        start = raw.find("[")
        end   = raw.rfind("]") + 1
        if start != -1 and end > start:
            raw = raw[start:end]
        raw = re.sub(r'//[^\n]*', '', raw)
        raw = re.sub(r',\s*([\]}])', r'\1', raw)
        ranking_list = json.loads(raw)
        return {item["nombre"]: item for item in ranking_list}
    except Exception as e:
        print(f"  [WARN] No se pudo rankear: {e}")
        return {}


# ---------------------------------------------------------------------------
# Proponer al universo — guarda pendientes, nunca toca companies.json solo
# ---------------------------------------------------------------------------
PENDING_FILE = BASE_DIR / "pending_universe.json"


def propose_to_universe(suggestions: list, companies_json: Path) -> list:
    """
    Filtra sugerencias con moat_strength='wide' que no estén ya en el universo
    y las guarda en pending_universe.json para confirmación manual.
    NO modifica companies.json nunca.
    """
    with open(companies_json, encoding="utf-8") as f:
        universe = json.load(f)

    existing_names   = {c["name"].lower() for c in universe}
    existing_tickers = {c["ticker"].lower() for c in universe}

    # Cargar pendientes ya existentes para no duplicar
    existing_pending = []
    if PENDING_FILE.exists():
        with open(PENDING_FILE, encoding="utf-8") as f:
            existing_pending = json.load(f)
    pending_names = {p["nombre"].lower() for p in existing_pending}

    proposed = []
    for s in suggestions:
        nombre   = s.get("nombre", "").strip()
        ticker   = s.get("ticker_yahoo", "").strip()
        moat     = s.get("moat_type", "none")
        strength = s.get("moat_strength", "").lower()

        if moat == "none":                          continue
        if strength != "wide":                      continue
        if nombre.lower() in existing_names:        continue
        if ticker.lower() in existing_tickers:      continue
        if nombre.lower() in pending_names:         continue

        proposed.append(s)
        pending_names.add(nombre.lower())

    if proposed:
        all_pending = existing_pending + proposed
        with open(PENDING_FILE, "w", encoding="utf-8") as f:
            json.dump(all_pending, f, ensure_ascii=False, indent=2)

    return proposed


def _infer_exchange(yf_ticker: str) -> str:
    """Infiere el exchange desde el sufijo del ticker Yahoo Finance."""
    suffix_to_exchange = {v: k for k, v in EXCHANGE_SUFFIX.items() if v}
    for suffix, exchange in suffix_to_exchange.items():
        if yf_ticker.endswith(suffix):
            return exchange
    return "XNAS"  # default US


# ---------------------------------------------------------------------------
# Construir geografía desde exchange
# ---------------------------------------------------------------------------
EXCHANGE_GEO = {
    "XLON": "Europa", "XETR": "Europa", "XPAR": "Europa", "XAMS": "Europa",
    "XSWX": "Europa", "XMIL": "Europa", "XMAD": "Europa",
    "XNAS": "US", "XNYS": "US",
    "XTKS": "Asia", "XHKG": "Asia", "XASX": "Asia",
    "XBOM": "Asia", "XNSE": "Asia", "XKRX": "Asia", "XTAI": "Asia",
    "XSHG": "Asia", "XSHE": "Asia",
    "BVMF": "Latam",
}


# ---------------------------------------------------------------------------
# FIX 2 — High conviction filter
# ---------------------------------------------------------------------------
def is_high_conviction(r: dict) -> bool:
    moat = r.get("moat", {})
    return (
        moat.get("vale_investigar") is True
        and moat.get("conviction_score", 0) >= 7
        and moat.get("moat_strength") == "wide"
        and moat.get("reversibilidad") in ("gestion", "ciclo")
    )


# ---------------------------------------------------------------------------
# FIX 3 — Timing filter
# ---------------------------------------------------------------------------
def is_good_timing(r: dict) -> bool:
    # Always show if conviction is 9 or 10 — urgent regardless of timing
    if r.get("moat", {}).get("conviction_score", 0) >= 9:
        return True

    tendencia = r.get("tendencia", "Nueva")

    # Nueva: first time seen — show it (might be a new opportunity)
    if tendencia == "Nueva":
        return True

    # Profundizando: derating getting worse — good, opportunity growing
    if tendencia == "Profundizando":
        return True

    # Recuperando: already recovering — timing may have passed, don't alert
    if tendencia == "Recuperando":
        return False

    # Estable: show only if it's been stable for less than 3 weeks
    ticker = r.get("yf_ticker", "")
    tracking_file = BASE_DIR / "tracking.json"
    if tracking_file.exists():
        with open(tracking_file) as f:
            tracking = json.load(f)
        entry = tracking.get(ticker, {})
        entries = entry.get("entries", [])
        if len(entries) >= 3:
            return False  # stable for 3+ weeks, skip
    return True


# ---------------------------------------------------------------------------
# Generar Excel — with IMPROVEMENTS 1, 2, 3
# ---------------------------------------------------------------------------
def build_excel(results: list, suggestions: list, output_path: Path):
    # FIX 1 — separate universe vs global
    universe_results = [r for r in results if r.get("source") != "fmp_global"]
    global_results   = [r for r in results if r.get("source") == "fmp_global"]

    wb = openpyxl.Workbook()

    # ---- Colores ----
    GREEN_DARK  = PatternFill("solid", fgColor="1A7A3C")
    GREEN_LIGHT = PatternFill("solid", fgColor="D6F0E0")
    RED_LIGHT   = PatternFill("solid", fgColor="FADADD")
    YELLOW      = PatternFill("solid", fgColor="FFF3CD")
    GRAY        = PatternFill("solid", fgColor="F2F2F2")
    HEADER_FILL = PatternFill("solid", fgColor="1A2B4A")

    white_bold = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    normal_font = Font(name="Calibri", size=10)
    bold_font = Font(bold=True, name="Calibri", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ============================================================
    # Pestaña 1 — Screener Results
    # ============================================================
    ws = wb.active
    ws.title = "Screener Results"

    headers = [
        "!", "Nombre", "Ticker", "Geografía", "Sector",
        "Caída_máximos_%", "P/E_actual", "P/E_compresión_%",
        "EV/EBITDA_actual", "EV/EBITDA_descuento_%",
        "Derating", "Moat_type", "Moat_strength", "Moat_intact",
        "Narrativa_negativa", "Reversibilidad", "Vale_investigar",
        "Conviction", "Conviction_razon",
        "Tendencia",
        "Razón", "Por qué es top", "Comparable", "Fuente", "Fecha_análisis"
    ]

    ws.row_dimensions[1].height = 28
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = white_bold
        cell.alignment = center
        cell.border = border

    col_widths = [4, 28, 14, 12, 22, 16, 12, 18, 16, 18, 10, 20, 16, 14, 45, 16, 16, 12, 45, 14, 50, 40, 20, 16, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    STAR_FILL = PatternFill("solid", fgColor="FFF0B0")  # amarillo dorado para destacadas

    for row_idx, r in enumerate(universe_results, 2):
        moat      = r.get("moat", {})
        sig       = r.get("signals", {})
        vale      = moat.get("vale_investigar")
        rev       = moat.get("reversibilidad", "")
        derating_pass = r.get("derating_pass", False)
        destacada = r.get("destacada", False)

        estrella = "★" if destacada else ""

        row_data = [
            estrella,
            r.get("name", ""),
            r.get("yf_ticker", ""),
            r.get("geo", ""),
            r.get("sector", ""),
            sig.get("drop_from_high_pct", ""),
            sig.get("pe_current", ""),
            sig.get("pe_compression_pct", ""),
            sig.get("ev_ebitda_current", ""),
            sig.get("ev_ebitda_discount_pct", ""),
            "SI" if derating_pass else "NO",
            moat.get("moat_type", "") if derating_pass else "",
            moat.get("moat_strength", "") if derating_pass else "",
            ("Si" if moat.get("moat_intact") else "No") if derating_pass and "moat_intact" in moat else "",
            moat.get("narrativa_negativa", "") if derating_pass else "",
            rev if derating_pass else "",
            ("Si" if vale else "No") if derating_pass and vale is not None else "",
            moat.get("conviction_score", "") if derating_pass else "",
            moat.get("conviction_razon", "") if derating_pass else "",
            r.get("tendencia", "") if derating_pass else "",
            moat.get("razon", "") if derating_pass else "",
            r.get("razon_prioridad", "") if derating_pass and vale else "",
            moat.get("comparable") or "" if derating_pass else "",
            "Global FMP" if r.get("source") == "fmp_global" else "Universo",
            date.today().strftime("%Y-%m-%d"),
        ]

        # Color de fila
        if destacada:
            row_fill = STAR_FILL
        elif not derating_pass:
            row_fill = GRAY
        elif vale is True:
            if rev in ("gestion", "ciclo"):
                row_fill = GREEN_LIGHT
            else:
                row_fill = YELLOW
        else:
            row_fill = RED_LIGHT

        left_cols = {2, 15, 19, 21, 22}
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            cell.font = bold_font if destacada else normal_font
            cell.alignment = left if col_idx in left_cols else center
            cell.border = border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(universe_results)+1}"

    # ============================================================
    # Pestaña 2 — Candidatas Globales (FMP)
    # ============================================================
    ws_global = wb.create_sheet("Candidatas Globales")
    ws_global.row_dimensions[1].height = 28
    for col_idx, h in enumerate(headers, 1):
        cell = ws_global.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = white_bold
        cell.alignment = center
        cell.border = border
    for i, w in enumerate(col_widths, 1):
        ws_global.column_dimensions[get_column_letter(i)].width = w

    for row_idx, r in enumerate(global_results, 2):
        moat      = r.get("moat", {})
        sig       = r.get("signals", {})
        vale      = moat.get("vale_investigar")
        rev       = moat.get("reversibilidad", "")
        derating_pass = r.get("derating_pass", False)
        destacada = r.get("destacada", False)
        estrella = "★" if destacada else ""
        row_data = [
            estrella, r.get("name", ""), r.get("yf_ticker", ""), r.get("geo", ""), r.get("sector", ""),
            sig.get("drop_from_high_pct", ""), sig.get("pe_current", ""), sig.get("pe_compression_pct", ""),
            sig.get("ev_ebitda_current", ""), sig.get("ev_ebitda_discount_pct", ""),
            "SI" if derating_pass else "NO",
            moat.get("moat_type", "") if derating_pass else "",
            moat.get("moat_strength", "") if derating_pass else "",
            ("Si" if moat.get("moat_intact") else "No") if derating_pass and "moat_intact" in moat else "",
            moat.get("narrativa_negativa", "") if derating_pass else "",
            rev if derating_pass else "",
            ("Si" if vale else "No") if derating_pass and vale is not None else "",
            moat.get("conviction_score", "") if derating_pass else "",
            moat.get("conviction_razon", "") if derating_pass else "",
            r.get("tendencia", "") if derating_pass else "",
            moat.get("razon", "") if derating_pass else "",
            r.get("razon_prioridad", "") if derating_pass and vale else "",
            moat.get("comparable") or "" if derating_pass else "",
            "Global FMP",
            date.today().strftime("%Y-%m-%d"),
        ]
        if not derating_pass:
            row_fill = GRAY
        elif vale is True:
            row_fill = GREEN_LIGHT if rev in ("gestion", "ciclo") else YELLOW
        else:
            row_fill = RED_LIGHT
        left_cols = {2, 15, 19, 21, 22}
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_global.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            cell.font = normal_font
            cell.alignment = left if col_idx in left_cols else center
            cell.border = border

    ws_global.freeze_panes = "A2"
    if global_results:
        ws_global.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(global_results)+1}"

    # ============================================================
    # Pestaña 3 — Sugerencias Claude
    # ============================================================
    ws2 = wb.create_sheet("Sugerencias Claude")
    sug_headers = ["Nombre", "Ticker Yahoo", "Geografía", "Sector", "Moat_type", "Por qué añadir"]
    ws2.row_dimensions[1].height = 28

    for col_idx, h in enumerate(sug_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = white_bold
        cell.alignment = center
        cell.border = border

    sug_widths = [28, 14, 12, 22, 20, 55]
    for i, w in enumerate(sug_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    for row_idx, s in enumerate(suggestions, 2):
        row_data = [
            s.get("nombre", ""),
            s.get("ticker_yahoo", ""),
            s.get("geografia", ""),
            s.get("sector", ""),
            s.get("moat_type", ""),
            s.get("por_que_añadir", ""),
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = GREEN_LIGHT
            cell.font = normal_font
            cell.alignment = left
            cell.border = border

    ws2.freeze_panes = "A2"

    wb.save(output_path)
    print(f"\n[OK] Excel guardado en: {output_path}")


# ---------------------------------------------------------------------------
# Email — with IMPROVEMENTS 1, 3, 4, 5
# ---------------------------------------------------------------------------
SMTP_HOST = os.environ.get("SMTP_HOST", "smtp.office365.com")
SMTP_PORT = int(os.environ.get("SMTP_PORT") or "587")
SMTP_USER = os.environ.get("SMTP_USER", "acb@estelacapital.es")
SMTP_PASS = os.environ.get("SMTP_PASS", "")
EMAIL_TO  = os.environ.get("EMAIL_TO", "acb@estelacapital.es")


def _proposed_section(proposed: list, perplexity_key: str = "") -> str:
    if not proposed:
        return ""
    cards_html = ""
    for s in proposed:
        # Remap _moat/_signals keys to match _company_card() expectations
        r = {
            "name":      s.get("nombre", ""),
            "ticker":    s.get("ticker_yahoo", ""),
            "geo":       s.get("geografia", ""),
            "sector":    s.get("sector", "") or s.get("geografia", ""),
            "moat":      s.get("_moat", {}),
            "signals":   s.get("_signals", {}),
            "destacada": False,
            "tendencia": "",
        }
        cards_html += _company_card(r, perplexity_key=perplexity_key, is_global=False)
    return f"""
      <div style="margin-top:32px">
        <div style="padding:16px 32px 8px;background:#1A2B4A;border-radius:8px 8px 0 0">
          <h2 style="color:#fff;margin:0;font-size:18px">&#128269; Propuestas para ampliar el universo</h2>
          <p style="color:#93c5fd;font-size:12px;margin:4px 0 0">
            {len(proposed)} empresa(s) descubiertas por IA con moat <strong>wide</strong>, derating real y narrativa reversible.
            Para aprobar o rechazar ejecuta: <code style="background:rgba(255,255,255,0.15);padding:2px 6px;border-radius:3px">python confirm.py</code>
          </p>
        </div>
        <div style="padding:16px 32px 24px;background:#F0F7FF;border:2px solid #2E86AB;border-top:none;border-radius:0 0 8px 8px">
          {cards_html}
        </div>
      </div>"""


def get_insider_activity(ticker: str, finnhub_key: str) -> str:
    """Returns a short string describing recent insider transactions, or empty string."""
    if not finnhub_key or not ticker:
        return ""
    try:
        import urllib.request as _ureq, json as _json, time as _t
        from datetime import date, timedelta
        since = (date.today() - timedelta(days=90)).isoformat()
        url = (f"https://finnhub.io/api/v1/stock/insider-transactions"
               f"?symbol={ticker}&from={since}&token={finnhub_key}")
        with _ureq.urlopen(url, timeout=8) as r:
            data = _json.loads(r.read().decode())
        txns = data.get("data") or []
        buys  = [t for t in txns if t.get("transactionType") in ("P - Purchase",) and t.get("share", 0) > 0]
        sells = [t for t in txns if t.get("transactionType") in ("S - Sale",) and t.get("share", 0) > 0]
        parts = []
        if buys:
            total_buy = sum(t.get("share", 0) for t in buys)
            parts.append(f"&#128200; {len(buys)} compra(s) insider ({total_buy:,.0f} acc.)")
        if sells:
            total_sell = sum(t.get("share", 0) for t in sells)
            parts.append(f"&#128201; {len(sells)} venta(s) insider ({total_sell:,.0f} acc.)")
        return " · ".join(parts) if parts else ""
    except Exception:
        return ""


def _price_sparkline_b64(signals: dict) -> str:
    """Generate a base64 PNG sparkline of 12-month price with derating zone highlighted."""
    history = signals.get("price_history")
    if not history or len(history) < 10:
        return ""
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import io, base64

        prices = history
        high_52w = signals.get("high_52w")
        current  = signals.get("current_price")
        p1m      = signals.get("price_1m_ago")

        fig, ax = plt.subplots(figsize=(4.5, 1.5), dpi=90)
        fig.patch.set_facecolor("#F8FAFC")
        ax.set_facecolor("#F8FAFC")

        x = list(range(len(prices)))

        # Derating zone shading (between 52w high and line)
        if high_52w:
            ax.fill_between(x, [high_52w] * len(x), prices,
                            where=[p < high_52w for p in prices],
                            color="#CC0000", alpha=0.08, zorder=1)
            ax.axhline(y=high_52w, color="#CC0000", linewidth=0.8,
                       linestyle="--", alpha=0.6, zorder=2)

        # Price line
        ax.plot(x, prices, color="#1A2B4A", linewidth=1.8, zorder=3)

        # 1-month ago marker
        if p1m and len(prices) >= 22:
            idx_1m = len(prices) - 22
            ax.plot(idx_1m, p1m, "o", color="#2E86AB", markersize=4, zorder=5)

        # Current price dot
        ax.plot(len(prices) - 1, prices[-1], "o", color="#CC0000", markersize=5, zorder=5)

        # Min/max y range with padding
        mn, mx = min(prices), max(prices)
        pad = (mx - mn) * 0.15 if mx > mn else mx * 0.05
        ax.set_ylim(mn - pad, mx + pad)
        ax.set_xlim(-3, len(prices) + 3)

        # Labels
        currency = signals.get("currency", "")
        if current:
            ax.annotate(f"{current:.1f}", xy=(len(prices) - 1, current),
                        xytext=(5, 0), textcoords="offset points",
                        fontsize=7, color="#CC0000", va="center", fontweight="bold")
        if high_52w:
            ax.annotate(f"Máx {high_52w:.1f}", xy=(0, high_52w),
                        xytext=(3, 4), textcoords="offset points",
                        fontsize=6, color="#CC0000", alpha=0.75)

        # Clean axes
        for spine in ax.spines.values():
            spine.set_visible(False)
        ax.tick_params(left=False, labelleft=False, bottom=False, labelbottom=False)
        ax.grid(axis="y", color="#e0e6f0", linewidth=0.5, alpha=0.7)

        plt.tight_layout(pad=0.3)
        buf = io.BytesIO()
        fig.savefig(buf, format="png", bbox_inches="tight", dpi=90,
                    facecolor="#F8FAFC", edgecolor="none")
        plt.close(fig)
        buf.seek(0)
        return base64.b64encode(buf.read()).decode("utf-8")
    except Exception:
        return ""


def _company_card(r: dict, perplexity_key: str = "", is_global: bool = False) -> str:
    """Rich HTML card for a single company — universe or global."""
    moat      = r.get("moat", {})
    sig       = r.get("signals", {})
    destacada = r.get("destacada", False)
    tendencia = r.get("tendencia", "")

    # Colors
    rev_color = {"gestion": "#1A7A3C", "ciclo": "#2E86AB", "mixto": "#E07B00", "estructural": "#CC0000"}.get(moat.get("reversibilidad", ""), "#555")
    tend_color = {"Nueva": "#7C3AED", "Profundizando": "#CC0000", "Recuperando": "#1A7A3C", "Estable": "#555"}.get(tendencia, "#555")
    strength_color = {"wide": "#1A7A3C", "narrow": "#E07B00", "none": "#CC0000"}.get(moat.get("moat_strength", ""), "#888")
    card_bg   = "#FFFBEB" if destacada else ("#F0F7FF" if is_global else "#FAFBFC")
    border_l  = "#F0A500" if destacada else ("#2E86AB" if is_global else "#1A2B4A")

    # Header badges
    star_badge = '<span style="background:#F0A500;color:#fff;padding:2px 8px;border-radius:10px;font-size:11px;margin-right:6px;font-weight:700">&#9733; DESTACADA</span>' if destacada else ""
    tend_badge = f'<span style="color:{tend_color};font-size:11px;font-weight:600;border:1px solid {tend_color};padding:1px 7px;border-radius:10px;margin-left:8px">{tendencia}</span>' if tendencia else ""
    global_badge = '<span style="background:#2E86AB;color:#fff;padding:2px 7px;border-radius:10px;font-size:10px;margin-left:6px">GLOBAL</span>' if is_global else ""
    intact_badge = ""
    if moat.get("moat_intact") is True:
        intact_badge = '<span style="background:#1A7A3C;color:#fff;padding:2px 6px;border-radius:8px;font-size:10px;margin-left:6px">moat intacto ✓</span>'
    elif moat.get("moat_intact") is False:
        intact_badge = '<span style="background:#CC0000;color:#fff;padding:2px 6px;border-radius:8px;font-size:10px;margin-left:6px">moat deteriorado ✗</span>'

    # Ticker + geo + sector (from company obj or signals)
    ticker  = r.get("ticker", "")
    geo     = r.get("geo", "")
    sector  = r.get("sector", "") or sig.get("sector", "")
    meta_parts = [p for p in [ticker, geo, sector] if p]
    meta_str = " · ".join(meta_parts)

    # Moat strength label
    strength = moat.get("moat_strength", "")
    strength_html = f'<span style="color:{strength_color};font-weight:700">{strength}</span>' if strength else ""

    # Key metrics
    drop         = sig.get("drop_from_high_pct", "")
    cur_price    = sig.get("current_price")
    price_1m     = sig.get("price_1m_ago")
    high_52w     = sig.get("high_52w")
    currency     = sig.get("currency", "")
    pe_cur       = sig.get("pe_current")
    pe_mean      = sig.get("pe_mean_5y")
    pe_comp      = sig.get("pe_compression_pct")
    ev_cur       = sig.get("ev_ebitda_current")
    ev_disc      = sig.get("ev_ebitda_discount_pct")
    roe          = sig.get("roe")
    gross_margin = sig.get("gross_margin")
    rev_growth   = sig.get("revenue_growth_1y")
    fcf_yield    = sig.get("fcf_yield")
    div_yield    = sig.get("dividend_yield")
    beta_val     = sig.get("beta")
    nd_ebitda    = sig.get("net_debt_ebitda")
    analyst_tgt  = sig.get("analyst_target")
    analyst_up   = sig.get("analyst_upside_pct")
    analyst_cnt  = sig.get("analyst_count")
    rec          = sig.get("recommendation", "")
    description  = sig.get("description", "")

    # Price vs 1 month ago — colored
    price_1m_html = ""
    if cur_price is not None and price_1m is not None:
        perf_1m = round((cur_price - price_1m) / price_1m * 100, 1)
        perf_color = "#1A7A3C" if perf_1m >= 0 else "#CC0000"
        perf_arrow = "▲" if perf_1m >= 0 else "▼"
        price_1m_html = f'<span style="color:{perf_color};font-weight:700;font-size:13px">{perf_arrow} {abs(perf_1m)}% vs hace 1 mes</span>'

    # Price row
    price_parts = []
    if cur_price is not None:
        price_parts.append(f'<strong>Precio:</strong> <span style="font-weight:700;font-size:14px">{cur_price} {currency}</span>')
    if price_1m_html:
        price_parts.append(price_1m_html)
    if high_52w is not None:
        price_parts.append(f'<strong>Máx. 52s:</strong> {high_52w} {currency}')
    price_html = f'<div style="font-size:12px;color:#555;margin-bottom:6px">{" &nbsp;·&nbsp; ".join(price_parts)}</div>' if price_parts else ""

    # Sparkline chart
    sparkline_b64 = _price_sparkline_b64(sig)
    sparkline_html = (
        f'<div style="margin:8px 0 4px">'
        f'<img src="data:image/png;base64,{sparkline_b64}" '
        f'style="width:100%;max-width:420px;height:auto;border-radius:4px;display:block" '
        f'alt="Precio 12 meses"/>'
        f'<div style="font-size:9px;color:#aaa;margin-top:1px">'
        f'&#9632; precio actual &nbsp;&#9588;&#9588; máximo 52 semanas &nbsp;&#9679; hace 1 mes</div>'
        f'</div>'
    ) if sparkline_b64 else ""

    # PE row
    pe_parts = []
    if pe_cur is not None:
        pe_parts.append(f'<strong>P/E:</strong> {pe_cur}x')
    if pe_mean is not None:
        pe_parts.append(f'<strong>P/E medio 5a:</strong> {pe_mean}x')
    if pe_comp is not None:
        pe_color_val = "#1A7A3C" if pe_comp > 0 else "#CC0000"
        pe_parts.append(f'<span style="color:{pe_color_val}"><strong>Compresión:</strong> {pe_comp:+.0f}%</span>')
    pe_html = f'<div style="font-size:12px;color:#555;margin-bottom:6px">{" &nbsp;·&nbsp; ".join(pe_parts)}</div>' if pe_parts else ""

    # Quality metrics row
    qual_parts = []
    if roe is not None:
        qual_parts.append(f'ROE: {roe}%')
    if gross_margin is not None:
        qual_parts.append(f'Margen bruto: {gross_margin}%')
    if fcf_yield is not None:
        qual_parts.append(f'FCF yield: {fcf_yield}%')
    if rev_growth is not None:
        rg_color = "#1A7A3C" if rev_growth >= 0 else "#CC0000"
        qual_parts.append(f'<span style="color:{rg_color}">Crecimiento: {rev_growth:+.1f}%</span>')
    if nd_ebitda is not None:
        nd_color = "#CC0000" if nd_ebitda > 3 else ("#E07B00" if nd_ebitda > 2 else "#1A7A3C")
        qual_parts.append(f'<span style="color:{nd_color}">Deuda/EBITDA: {nd_ebitda}x</span>')
    if div_yield is not None:
        qual_parts.append(f'Dividendo: {div_yield}%')
    if beta_val is not None:
        qual_parts.append(f'Beta: {beta_val}')
    qual_html = f'<div style="font-size:11px;color:#666;margin-bottom:6px">{" &nbsp;·&nbsp; ".join(qual_parts)}</div>' if qual_parts else ""

    # Analyst consensus row
    analyst_html = ""
    if analyst_tgt is not None:
        up_color = "#1A7A3C" if (analyst_up or 0) > 0 else "#CC0000"
        up_str = f' <span style="color:{up_color};font-weight:700">({analyst_up:+.1f}%)</span>' if analyst_up is not None else ""
        rec_str = f" · Consenso: <strong>{rec}</strong>" if rec else ""
        cnt_str = f" · {analyst_cnt} analistas" if analyst_cnt else ""
        analyst_html = f'<div style="font-size:11px;color:#555;margin-bottom:4px">&#127919; Objetivo analistas: <strong>{analyst_tgt} {currency}</strong>{up_str}{rec_str}{cnt_str}</div>'

    metrics_parts = [f'<span style="color:#CC0000;font-weight:700;font-size:15px">&#8595; {drop}%</span> desde máximos']
    if ev_cur is not None:
        ev_str = f"EV/EBITDA {ev_cur}x"
        if ev_disc is not None:
            ev_str += f" ({ev_disc:+.1f}% vs hist.)"
        metrics_parts.append(f'<span style="color:#555">{ev_str}</span>')
    metrics_html = ' &nbsp;|&nbsp; '.join(metrics_parts)

    # Business description
    desc_html = f"""
      <div style="margin-top:8px;padding:8px 12px;background:#f9f9f9;border-left:3px solid #ccc;border-radius:0 4px 4px 0">
        <div style="font-size:11px;color:#888;font-weight:600;text-transform:uppercase;letter-spacing:.5px;margin-bottom:2px">&#127970; Descripción del negocio</div>
        <div style="font-size:12px;color:#555;line-height:1.5">{description}</div>
      </div>""" if description else ""

    # Moat line
    moat_type = moat.get("moat_type", "")
    moat_line = f"{moat_type}"
    if strength:
        moat_line += f" &nbsp;·&nbsp; {strength_html}"
    if intact_badge:
        moat_line += f" &nbsp;{intact_badge}"

    # Narrativa negativa
    narrativa = moat.get("narrativa_negativa", "")
    narrativa_html = f"""
      <div style="margin-top:10px;padding:8px 12px;background:#fff5f5;border-left:3px solid #CC0000;border-radius:0 4px 4px 0">
        <div style="font-size:11px;color:#888;font-weight:600;text-transform:uppercase;letter-spacing:.5px;margin-bottom:2px">&#128200; Qué teme el mercado</div>
        <div style="font-size:12px;color:#444;line-height:1.5">{narrativa}</div>
      </div>""" if narrativa else ""

    # Razón / tesis
    razon = moat.get("razon", "")
    razon_html = f"""
      <div style="margin-top:8px;padding:8px 12px;background:#f0fff4;border-left:3px solid #1A7A3C;border-radius:0 4px 4px 0">
        <div style="font-size:11px;color:#888;font-weight:600;text-transform:uppercase;letter-spacing:.5px;margin-bottom:2px">&#128161; Tesis Estela</div>
        <div style="font-size:12px;color:#1A3A2A;line-height:1.5">{razon}</div>
      </div>""" if razon else ""

    # Convicción
    conviction = moat.get("conviction_score", "")
    conv_razon = moat.get("conviction_razon", "")
    conv_color = "#1A7A3C" if conviction and int(conviction) >= 7 else ("#E07B00" if conviction and int(conviction) >= 5 else "#CC0000")
    conviction_html = f"""
      <div style="margin-top:8px;padding:8px 12px;background:#f8f6ff;border-left:3px solid #7C3AED;border-radius:0 4px 4px 0">
        <div style="font-size:11px;color:#888;font-weight:600;text-transform:uppercase;letter-spacing:.5px;margin-bottom:2px">&#11088; Convicción</div>
        <div style="font-size:12px;color:#333;line-height:1.5"><span style="font-size:18px;font-weight:700;color:{conv_color}">{conviction}</span><span style="color:#888">/10</span>&nbsp;&nbsp;{conv_razon}</div>
      </div>""" if conviction else ""

    # Comparable
    comparable = moat.get("comparable")
    comparable_html = f'<div style="margin-top:6px;font-size:11px;color:#888">&#128257; Similar a: <em>{comparable}</em></div>' if comparable and comparable != "null" else ""

    # Catalysts
    catalysts = moat.get("catalysts") or []
    if catalysts and isinstance(catalysts, list) and len(catalysts) > 0:
        cat_items = "".join(f'<li style="margin-bottom:3px">{c}</li>' for c in catalysts if c)
        catalysts_html = f"""
      <div style="margin-top:8px;padding:8px 12px;background:#f0f9ff;border-left:3px solid #2E86AB;border-radius:0 4px 4px 0">
        <div style="font-size:11px;color:#888;font-weight:600;text-transform:uppercase;letter-spacing:.5px;margin-bottom:4px">&#128640; Catalizadores</div>
        <ul style="margin:0;padding-left:16px;font-size:12px;color:#1A3A5C;line-height:1.6">{cat_items}</ul>
      </div>"""
    else:
        catalysts_html = ""

    # Risks
    risks = moat.get("risks") or []
    if risks and isinstance(risks, list) and len(risks) > 0:
        risk_items = "".join(f'<li style="margin-bottom:3px">{rk}</li>' for rk in risks if rk)
        risks_html = f"""
      <div style="margin-top:8px;padding:8px 12px;background:#fff5f5;border-left:3px solid #E07B00;border-radius:0 4px 4px 0">
        <div style="font-size:11px;color:#888;font-weight:600;text-transform:uppercase;letter-spacing:.5px;margin-bottom:4px">&#9888; Riesgos</div>
        <ul style="margin:0;padding-left:16px;font-size:12px;color:#5C2A00;line-height:1.6">{risk_items}</ul>
      </div>"""
    else:
        risks_html = ""

    # Entry price
    entry_comment = moat.get("entry_price_comment", "")
    entry_html = f"""
      <div style="margin-top:8px;padding:8px 12px;background:#f6fff6;border-left:3px solid #1A7A3C;border-radius:0 4px 4px 0">
        <div style="font-size:11px;color:#888;font-weight:600;text-transform:uppercase;letter-spacing:.5px;margin-bottom:2px">&#127919; Precio de entrada</div>
        <div style="font-size:12px;color:#1A3A2A;line-height:1.5">{entry_comment}</div>
      </div>""" if entry_comment else ""

    # Reversibilidad badge
    rev = moat.get("reversibilidad", "")
    rev_html = f'<span style="color:{rev_color};font-weight:600;font-size:12px;border:1px solid {rev_color};padding:2px 8px;border-radius:10px">{rev}</span>' if rev else ""

    # Razon prioridad (destacada)
    razon_top = r.get("razon_prioridad", "")
    razon_top_html = f'<div style="margin-top:6px;font-size:11px;color:#9B7000;font-style:italic">&#128204; {razon_top}</div>' if razon_top and destacada else ""

    # News via Perplexity
    news_html = ""
    if perplexity_key:
        news = get_news_context(r.get("name", ""), perplexity_key)
        if news:
            news_html = f"""
      <div style="margin-top:8px;padding:8px 12px;background:#f5f5f5;border-left:3px solid #aaa;border-radius:0 4px 4px 0">
        <div style="font-size:11px;color:#888;font-weight:600;text-transform:uppercase;letter-spacing:.5px;margin-bottom:2px">&#128240; Noticias recientes</div>
        <div style="font-size:11px;color:#555;line-height:1.6;white-space:pre-wrap">{news}</div>
      </div>"""

    # Vault
    vault_html = ""
    vault_found, vault_excerpt = get_vault_excerpt(r.get("name", ""))
    if vault_found:
        safe_excerpt = vault_excerpt.replace("<", "&lt;").replace(">", "&gt;")
        vault_html = f"""
      <div style="margin-top:8px;padding:8px 12px;background:#f0fff4;border-left:3px solid #1A7A3C;border-radius:0 4px 4px 0">
        <div style="font-size:11px;color:#1A7A3C;font-weight:600">&#128218; Tesis en vault</div>
        <div style="font-size:11px;color:#555;font-style:italic;margin-top:2px">{safe_excerpt}...</div>
      </div>"""

    # Insider buying (Finnhub) — only for universe companies with a clean ticker
    insider_html = ""
    finnhub_key = os.environ.get("FINNHUB_API_KEY", "")
    insider_str = get_insider_activity(r.get("ticker", ""), finnhub_key)
    if insider_str:
        insider_html = f'<div style="margin-top:6px;font-size:11px;color:#1A3A5C;font-weight:600">{insider_str}</div>'

    return f"""
    <div style="margin-bottom:20px;padding:18px 20px;background:{card_bg};border:1px solid #e0e6f0;border-left:4px solid {border_l};border-radius:6px">
      <div style="display:flex;align-items:center;flex-wrap:wrap;gap:4px;margin-bottom:4px">
        {star_badge}
        <span style="font-size:17px;font-weight:700;color:#1A2B4A">{r.get('name','')}</span>
        {global_badge}{tend_badge}
      </div>
      <div style="font-size:11px;color:#888;margin-bottom:6px">{meta_str}</div>
      {price_html}
      {sparkline_html}
      {analyst_html}
      {qual_html}
      {pe_html}
      <div style="font-size:12px;color:#555;margin-bottom:8px">
        <div style="font-size:13px;margin-bottom:4px">{metrics_html}</div>
        <strong>Moat:</strong> {moat_line} &nbsp;&nbsp; <strong>Reversibilidad:</strong> {rev_html}
      </div>
      {insider_html}
      {razon_top_html}
      {desc_html}
      {narrativa_html}
      {razon_html}
      {conviction_html}
      {catalysts_html}
      {risks_html}
      {entry_html}
      {comparable_html}
      {news_html}
      {vault_html}
    </div>"""


def _global_candidates_section(global_vale: list) -> str:
    """HTML section for global FMP candidates that pass vale_investigar."""
    if not global_vale:
        return ""
    cards = "".join(_company_card(r, perplexity_key="", is_global=True) for r in global_vale)
    return f"""
      <div style="margin-top:28px">
        <div style="background:#1A3A5C;padding:14px 24px;border-radius:8px 8px 0 0">
          <h3 style="color:#fff;margin:0;font-size:16px">&#127758; Oportunidades globales (fuera del universo)</h3>
          <p style="color:#9bb8d4;font-size:12px;margin:4px 0 0">{len(global_vale)} empresa(s) detectadas via FMP Global Scan con derating real y moat intacto</p>
        </div>
        <div style="border:1px solid #c7dff0;border-top:none;border-radius:0 0 8px 8px;padding:16px">{cards}</div>
      </div>"""


def _waiting_section(waiting: list) -> str:
    """HTML note for high-conviction companies skipped due to timing."""
    if not waiting:
        return ""
    names = ", ".join(r.get("name", "") for r in waiting)
    return f"""
      <div style="margin-top:16px;padding:12px 20px;background:#fff8e1;border:1px solid #ffe082;border-radius:6px">
        <p style="margin:0;font-size:12px;color:#7a6000">
          <strong>En seguimiento (esperando mejor timing):</strong> {names}
        </p>
      </div>"""


def send_email(output_path: Path, results: list, threshold: float, proposed: list = None,
               perplexity_key: str = "", global_candidates: list = None, cloud: bool = False):
    # FIX 1 — Split universe vs global
    universe_results = [r for r in results if r.get("source", "universo") != "fmp_global"]
    global_results   = global_candidates or [r for r in results if r.get("source") == "fmp_global"]
    candidates = [r for r in universe_results if r.get("derating_pass")]
    proposed = proposed or []

    # FIX 2+3 — Apply high conviction + timing filters for email
    universe_high_conv = [r for r in universe_results if is_high_conviction(r)]
    universe_email     = [r for r in universe_high_conv if is_good_timing(r)]
    universe_waiting   = [r for r in universe_high_conv if not is_good_timing(r)]

    global_high_conv = [r for r in global_results if is_high_conviction(r)]
    global_email     = [r for r in global_high_conv if is_good_timing(r)]
    global_waiting   = [r for r in global_high_conv if not is_good_timing(r)]

    # Legacy counts for stats block
    vale = [r for r in universe_results if r.get("moat", {}).get("vale_investigar") is True]
    global_vale = [r for r in global_results if r.get("moat", {}).get("vale_investigar") is True]

    # Build company cards for universe picks
    if universe_email:
        cards_html = "".join(_company_card(r, perplexity_key) for r in universe_email)
    else:
        cards_html = '<p style="color:#888;font-style:italic;padding:12px 0">Esta semana no hay candidatas del universo con convicción suficiente.</p>'

    global_stat_td = ""
    if global_email:
        global_stat_td = f"""
            <td style="width:12px"></td>
            <td style="padding:12px 16px;background:#fff;border-radius:6px;border:1px solid #e0e6f0;text-align:center">
              <div style="font-size:28px;font-weight:700;color:#2E86AB">{len(global_email)}</div>
              <div style="font-size:12px;color:#888;margin-top:2px">Candidatas globales FMP</div>
            </td>"""

    html = f"""
    <html><body style="font-family:Calibri,Arial,sans-serif;color:#1A2B4A;max-width:900px;margin:0 auto">
      <div style="background:#1A2B4A;padding:24px 32px;border-radius:8px 8px 0 0">
        <h2 style="color:#fff;margin:0;font-size:22px">Screener Moats — Estela Capital</h2>
        <p style="color:#aac;margin:4px 0 0">{date.today().strftime("%A, %d %B %Y")}</p>
      </div>
      <div style="background:#f7f9fc;padding:20px 32px;border:1px solid #e0e6f0;border-top:none">
        <table style="width:100%;border-collapse:collapse">
          <tr>
            <td style="padding:12px 16px;background:#fff;border-radius:6px;border:1px solid #e0e6f0;text-align:center">
              <div style="font-size:28px;font-weight:700;color:#1A2B4A">{len(universe_results)}</div>
              <div style="font-size:12px;color:#888;margin-top:2px">Empresas analizadas</div>
            </td>
            <td style="width:12px"></td>
            <td style="padding:12px 16px;background:#fff;border-radius:6px;border:1px solid #e0e6f0;text-align:center">
              <div style="font-size:28px;font-weight:700;color:#CC0000">{len(candidates)}</div>
              <div style="font-size:12px;color:#888;margin-top:2px">Con derating &gt;={threshold:.0f}%</div>
            </td>
            <td style="width:12px"></td>
            <td style="padding:12px 16px;background:#fff;border-radius:6px;border:1px solid #e0e6f0;text-align:center">
              <div style="font-size:28px;font-weight:700;color:#1A7A3C">{len(vale)}</div>
              <div style="font-size:12px;color:#888;margin-top:2px">Vale investigar</div>
            </td>{global_stat_td}
          </tr>
        </table>
      </div>
      <div style="padding:24px 32px;border:1px solid #e0e6f0;border-top:none">
        <h3 style="color:#1A2B4A;margin:0 0 4px">Candidatas del universo para investigar</h3>
        <p style="color:#888;font-size:12px;margin:0 0 16px">Empresas de tu universo con derating &ge;{threshold:.0f}% y moat intacto</p>
        {cards_html}
      </div>
      <div style="padding:16px 32px;background:#f7f9fc;border:1px solid #e0e6f0;border-top:none;border-radius:0 0 8px 8px">
        <p style="margin:0;font-size:12px;color:#999">Excel completo adjunto · Generado automaticamente por Screener Moats</p>
      {_proposed_section(proposed, perplexity_key=perplexity_key)}
      {_global_candidates_section(global_email)}
      {_waiting_section(universe_waiting + global_waiting)}
      </div>
    </body></html>"""

    high_conv = [r for r in results if is_high_conviction(r)]
    msg = MIMEMultipart()
    msg["From"]    = SMTP_USER
    msg["To"]      = EMAIL_TO
    msg["Subject"] = f"Screener Moats — {len(high_conv)} ideas de alta conviccion — {date.today().strftime('%d/%m/%Y')}"

    msg.attach(MIMEText(html, "html"))

    # Attach Excel
    with open(output_path, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f"attachment; filename={output_path.name}")
    msg.attach(part)

    # En modo cloud (GitHub Actions) usamos SMTP directamente — no hay Outlook
    if cloud:
        try:
            server = smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30)
            server.ehlo()
            server.starttls()
            server.login(SMTP_USER, SMTP_PASS)
            server.sendmail(SMTP_USER, EMAIL_TO, msg.as_string())
            server.quit()
            print(f"[OK] Email enviado a {EMAIL_TO} via SMTP (cloud)")
        except Exception as e:
            print(f"[ERROR] SMTP falló: {e}")
    else:
        try:
            import win32com.client
            outlook = win32com.client.Dispatch("Outlook.Application")
            outlook.Session.Logon()
            mail = outlook.CreateItem(0)
            mail.To = EMAIL_TO
            mail.Subject = msg["Subject"]
            mail.HTMLBody = html
            mail.Attachments.Add(str(output_path.resolve()))
            mail.Send()
            print(f"[OK] Email enviado a {EMAIL_TO} via Outlook")
        except Exception as e:
            print(f"[WARN] Outlook COM falló ({e}), intentando SMTP...")
            try:
                server = smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30)
                server.ehlo()
                server.starttls()
                server.login(SMTP_USER, SMTP_PASS)
                server.sendmail(SMTP_USER, EMAIL_TO, msg.as_string())
                server.quit()
                print(f"[OK] Email enviado a {EMAIL_TO} via SMTP")
            except Exception as e2:
                print(f"[ERROR] No se pudo enviar el email ni via Outlook ni SMTP: {e2}")


# ---------------------------------------------------------------------------
# Web dashboard (GitHub Pages)
# ---------------------------------------------------------------------------
def generate_web_dashboard(results: list, proposed: list, run_date: str) -> str:
    """Generate a rich interactive HTML dashboard for GitHub Pages."""
    all_companies = [r for r in results if r.get("moat", {}).get("vale_investigar")]

    def card_data(r):
        sig  = r.get("signals", {})
        moat = r.get("moat", {})
        history = sig.get("price_history") or []
        return {
            "name":        r.get("name", r.get("nombre", "")),
            "ticker":      r.get("ticker", ""),
            "sector":      r.get("sector", "") or sig.get("sector", ""),
            "geo":         r.get("geo", ""),
            "source":      r.get("source", "universo"),
            "drop":        sig.get("drop_from_high_pct", 0),
            "conviction":  moat.get("conviction_score", 0) or 0,
            "moat_type":   moat.get("moat_type", ""),
            "reversib":    moat.get("reversibilidad", ""),
            "razon":       moat.get("razon", ""),
            "price":       sig.get("current_price"),
            "high_52w":    sig.get("high_52w"),
            "currency":    sig.get("currency", ""),
            "roe":         sig.get("roe"),
            "gross_margin":sig.get("gross_margin"),
            "fcf_yield":   sig.get("fcf_yield"),
            "analyst_up":  sig.get("analyst_upside_pct"),
            "history":     history[-60:] if len(history) > 60 else history,  # last 60 days
            "catalysts":   moat.get("catalysts", []),
            "risks":       moat.get("risks", []),
            "entry":       moat.get("entry_price_comment", ""),
        }

    cards_json = json.dumps([card_data(r) for r in all_companies], ensure_ascii=False)
    proposed_json = json.dumps([card_data({
        "name": p.get("nombre",""), "ticker": p.get("ticker_yahoo",""),
        "sector": p.get("sector",""), "geo": p.get("geografia",""),
        "source": "propuesta", "signals": p.get("_signals",{}), "moat": p.get("_moat",{}),
    }) for p in proposed], ensure_ascii=False)

    return f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Estela Capital — Screener {run_date}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4/dist/chart.umd.min.js"></script>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
         background: #F0F4F8; color: #1A2B4A; }}
  header {{ background: #1A2B4A; color: #fff; padding: 20px 32px;
            display: flex; align-items: center; justify-content: space-between; }}
  header h1 {{ font-size: 22px; font-weight: 700; }}
  header span {{ font-size: 13px; opacity: .7; }}
  .filters {{ padding: 16px 32px; background: #fff; border-bottom: 1px solid #e0e6f0;
              display: flex; gap: 12px; flex-wrap: wrap; align-items: center; }}
  .filters label {{ font-size: 12px; color: #666; }}
  .filters select, .filters input {{ padding: 5px 10px; border: 1px solid #d0d8e8;
    border-radius: 6px; font-size: 13px; background: #fff; }}
  .grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(340px, 1fr));
           gap: 20px; padding: 24px 32px; }}
  .card {{ background: #fff; border-radius: 10px; border: 1px solid #e0e6f0;
           border-left: 4px solid #1A2B4A; overflow: hidden;
           transition: transform .15s, box-shadow .15s; }}
  .card:hover {{ transform: translateY(-2px); box-shadow: 0 6px 20px rgba(0,0,0,.1); }}
  .card.global {{ border-left-color: #2E86AB; }}
  .card.propuesta {{ border-left-color: #7C3AED; background: #faf8ff; }}
  .card-header {{ padding: 14px 16px 10px; }}
  .card-name {{ font-size: 16px; font-weight: 700; color: #1A2B4A; }}
  .card-meta {{ font-size: 11px; color: #888; margin-top: 2px; }}
  .badge {{ display: inline-block; padding: 1px 7px; border-radius: 10px;
            font-size: 10px; font-weight: 700; margin-left: 5px; }}
  .badge-global {{ background: #2E86AB; color: #fff; }}
  .badge-propuesta {{ background: #7C3AED; color: #fff; }}
  .chart-wrap {{ padding: 0 16px; height: 90px; position: relative; }}
  .card-body {{ padding: 10px 16px 14px; }}
  .drop {{ font-size: 20px; font-weight: 800; color: #CC0000; }}
  .metrics {{ display: flex; flex-wrap: wrap; gap: 6px; margin: 8px 0; }}
  .metric {{ background: #F0F4F8; border-radius: 5px; padding: 3px 8px;
             font-size: 11px; color: #444; }}
  .conviction {{ display: flex; align-items: center; gap: 6px; margin: 8px 0; }}
  .conv-bar {{ flex: 1; height: 6px; background: #e0e6f0; border-radius: 3px; }}
  .conv-fill {{ height: 100%; border-radius: 3px; transition: width .6s ease; }}
  .razon {{ font-size: 11px; color: #555; line-height: 1.5; margin-top: 6px;
            padding: 6px 10px; background: #f0fff4; border-left: 3px solid #1A7A3C;
            border-radius: 0 4px 4px 0; }}
  .section-title {{ font-size: 18px; font-weight: 700; color: #1A2B4A;
                    padding: 8px 32px 0; }}
  footer {{ text-align: center; padding: 24px; font-size: 11px; color: #aaa; }}
</style>
</head>
<body>
<header>
  <h1>&#127758; Estela Capital · Screener de Moats</h1>
  <span>Actualizado: {run_date}</span>
</header>

<div class="filters">
  <label>Filtrar: <input type="text" id="search" placeholder="empresa o ticker..." oninput="render()"/></label>
  <label>Fuente: <select id="srcFilter" onchange="render()">
    <option value="">Todas</option>
    <option value="universo">Universo</option>
    <option value="index_universe">Global</option>
    <option value="propuesta">Propuestas</option>
  </select></label>
  <label>Ordenar: <select id="sortBy" onchange="render()">
    <option value="conviction">Convicción</option>
    <option value="drop">Mayor caída</option>
    <option value="analyst_up">Upside analistas</option>
  </select></label>
  <span id="count" style="font-size:12px;color:#888;margin-left:auto"></span>
</div>

<div id="grid" class="grid"></div>
<footer>Generado automáticamente por Screener Moats · Estela Capital</footer>

<script>
const ALL = {cards_json};
const PROP = {proposed_json};
const DATA = [...ALL, ...PROP];
const charts = {{}};

function convColor(c) {{
  if (c >= 8) return "#1A7A3C";
  if (c >= 6) return "#E07B00";
  return "#CC0000";
}}

function makeCard(d, idx) {{
  const globalBadge = d.source !== "universo" && d.source !== "propuesta"
    ? '<span class="badge badge-global">GLOBAL</span>' : "";
  const propBadge = d.source === "propuesta"
    ? '<span class="badge badge-propuesta">PROPUESTA</span>' : "";
  const cls = d.source === "propuesta" ? "card propuesta"
            : d.source !== "universo"  ? "card global" : "card";
  const conv = d.conviction || 0;
  const metrics = [
    d.roe        ? `ROE: ${{d.roe}}%` : null,
    d.gross_margin ? `Margen: ${{d.gross_margin}}%` : null,
    d.fcf_yield  ? `FCF: ${{d.fcf_yield}}%` : null,
    d.analyst_up ? `Analistas: +${{d.analyst_up}}%` : null,
  ].filter(Boolean).map(m => `<span class="metric">${{m}}</span>`).join("");

  return `<div class="${{cls}}" data-idx="${{idx}}" data-src="${{d.source}}"
               data-name="${{d.name.toLowerCase()}}" data-ticker="${{d.ticker.toLowerCase()}}"
               data-conv="${{conv}}" data-drop="${{d.drop}}" data-up="${{d.analyst_up||0}}">
    <div class="card-header">
      <div class="card-name">${{d.name}}${{globalBadge}}${{propBadge}}</div>
      <div class="card-meta">${{d.ticker}} &nbsp;·&nbsp; ${{d.sector}} &nbsp;·&nbsp; ${{d.geo}}</div>
    </div>
    <div class="chart-wrap"><canvas id="chart-${{idx}}" height="90"></canvas></div>
    <div class="card-body">
      <div class="drop">&#8595; ${{d.drop}}% desde máximos</div>
      <div class="metrics">${{metrics}}</div>
      <div class="conviction">
        <span style="font-size:11px;color:#888">Convicción</span>
        <div class="conv-bar"><div class="conv-fill" style="width:${{conv*10}}%;background:${{convColor(conv)}}"></div></div>
        <span style="font-weight:700;color:${{convColor(conv)}}">${{conv}}/10</span>
      </div>
      ${{d.razon ? `<div class="razon">${{d.razon}}</div>` : ""}}
    </div>
  </div>`;
}}

function drawChart(d, idx) {{
  const canvas = document.getElementById("chart-" + idx);
  if (!canvas || !d.history || d.history.length < 5) return;
  if (charts[idx]) {{ charts[idx].destroy(); delete charts[idx]; }}
  const prices = d.history;
  const high = d.high_52w;
  charts[idx] = new Chart(canvas, {{
    type: "line",
    data: {{
      labels: prices.map((_, i) => i),
      datasets: [
        {{
          data: prices,
          borderColor: "#1A2B4A",
          borderWidth: 2,
          pointRadius: 0,
          fill: false,
          tension: 0.3,
        }},
        high ? {{
          data: prices.map(() => high),
          borderColor: "#CC0000",
          borderWidth: 1,
          borderDash: [4, 3],
          pointRadius: 0,
          fill: false,
        }} : null,
      ].filter(Boolean),
    }},
    options: {{
      animation: {{ duration: 800, easing: "easeInOutQuart" }},
      responsive: true,
      maintainAspectRatio: false,
      plugins: {{ legend: {{ display: false }}, tooltip: {{ enabled: false }} }},
      scales: {{
        x: {{ display: false }},
        y: {{ display: false, grace: "10%" }},
      }},
    }},
  }});
}}

function render() {{
  const q   = document.getElementById("search").value.toLowerCase();
  const src = document.getElementById("srcFilter").value;
  const srt = document.getElementById("sortBy").value;

  let filtered = DATA.filter(d => {{
    if (src && d.source !== src) return false;
    if (q && !d.name.toLowerCase().includes(q) && !d.ticker.toLowerCase().includes(q)) return false;
    return true;
  }});

  filtered.sort((a, b) => {{
    if (srt === "drop")     return b.drop - a.drop;
    if (srt === "analyst_up") return (b.analyst_up||0) - (a.analyst_up||0);
    return (b.conviction||0) - (a.conviction||0);
  }});

  document.getElementById("count").textContent = filtered.length + " empresas";
  const grid = document.getElementById("grid");
  grid.innerHTML = "";

  filtered.forEach((d, i) => {{
    const origIdx = DATA.indexOf(d);
    grid.insertAdjacentHTML("beforeend", makeCard(d, origIdx));
    requestAnimationFrame(() => drawChart(d, origIdx));
  }});
}}

render();
</script>
</body>
</html>"""


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Screener Moats — Estela Capital")
    parser.add_argument("--test", type=int, default=None, help="Limitar a N empresas (debug)")
    parser.add_argument("--threshold", type=float, default=DERATING_THRESHOLD, help="Umbral de derating (porcentaje)")
    parser.add_argument("--skip-moat", action="store_true", help="Solo calcular derating, sin llamar a Claude")
    parser.add_argument("--skip-global", action="store_true", help="Omitir scan global, solo universo Estela")
    parser.add_argument("--only-universe", choices=["portfolio", "watchlist", "universe", "all"], default="all")
    parser.add_argument("--cloud", action="store_true", help="Modo cloud (GitHub Actions): usa SMTP, sin Outlook COM")
    args = parser.parse_args()

    threshold = args.threshold

    # --- API Keys ---
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    gemini_key = os.environ.get("GEMINI_API_KEY", "")
    groq_key = os.environ.get("GROQ_API_KEY", "")
    perplexity_key = os.environ.get("PERPLEXITY_API_KEY", "")
    fmp_key = os.environ.get("FMP_API_KEY", "")

    # Elegir cliente AI: Anthropic si tiene saldo, Groq como fallback gratuito
    client = None
    if not args.skip_moat:
        if api_key and anthropic:
            try:
                _test = anthropic.Anthropic(api_key=api_key)
                _test.messages.create(model=CLAUDE_MODEL, max_tokens=5,
                    messages=[{"role":"user","content":"hi"}])
                client = _test
                print("[AI] Usando Claude (Anthropic)")
            except Exception as _e:
                print("[AI] Anthropic sin créditos — usando Groq")
                client = None
        if client is None and groq_key:
            client = GroqClient(groq_key)
            print("[AI] Usando Groq (Llama 3.3 70B — gratis)")
        if client is None:
            print("[WARN] Sin cliente AI — solo derating, sin análisis de moat")

    # --- Cargar empresas ---
    print(f"[1/4] Cargando universo desde {COMPANIES_JSON}...")
    with open(COMPANIES_JSON, encoding="utf-8") as f:
        all_companies = json.load(f)

    if args.only_universe != "all":
        all_companies = [c for c in all_companies if c.get("list") == args.only_universe]

    if args.test:
        all_companies = all_companies[:args.test]

    print(f"      {len(all_companies)} empresas en el universo")

    # --- Load persistent data ---
    ev_history = load_ev_history()
    tracking   = load_tracking()

    # --- PASO 1: Señal de derating ---
    print(f"\n[2/4] Calculando señales de derating (umbral: {threshold}%)...")
    results = []
    for i, company in enumerate(all_companies, 1):
        yf_ticker = resolve_ticker(company["name"], company["ticker"], company.get("exchange", ""))
        geo = EXCHANGE_GEO.get(company.get("exchange", ""), "Global")

        sys.stdout.write(f"  [{i:3d}/{len(all_companies)}] {company['name']:<35} {yf_ticker:<12} ")
        sys.stdout.flush()

        signals = get_derating_signals(yf_ticker, ev_history)
        passes = passes_derating_threshold(signals, threshold)

        if "error" in signals:
            sys.stdout.write(f"ERROR: {signals['error']}\n")
        elif passes:
            sys.stdout.write(f"DERATING {signals['drop_from_high_pct']}% caída\n")
        else:
            drop = signals.get("drop_from_high_pct", 0) or 0
            sys.stdout.write(f"ok (caída: {drop}%)\n")

        results.append({
            "name": company["name"],
            "ticker": company["ticker"],
            "yf_ticker": yf_ticker,
            "geo": geo,
            "sector": company.get("sector", ""),
            "signals": signals,
            "derating_pass": passes,
            "moat": {},
            "tendencia": "",
        })

    # Save updated EV history
    save_ev_history(ev_history)

    candidates = [r for r in results if r["derating_pass"]]
    print(f"\n      {len(candidates)} empresas pasan el umbral de derating")

    # --- PASO 2: Análisis de moats ---
    if not args.skip_moat and candidates:
        print(f"\n[3/4] Analizando moats con Claude ({CLAUDE_MODEL})...")
        for i, r in enumerate(candidates, 1):
            print(f"  [{i:2d}/{len(candidates)}] {r['name']}")
            moat = analyze_moat(client, r, r["signals"])
            r["moat"] = moat
            if "error" in moat:
                print(f"         [WARN] {moat['error']}")
            else:
                vale = "[SI] INVESTIGAR" if moat.get("vale_investigar") else "[NO] descartar"
                conviction = moat.get("conviction_score", "?")
                print(f"         {vale} | {moat.get('moat_type','')} | {moat.get('reversibilidad','')} | conviccion: {conviction}/10")

            # --- Improvement 3: Update tracking ---
            drop_pct = r["signals"].get("drop_from_high_pct", 0) or 0
            vale_bool = moat.get("vale_investigar", False)
            conviction_int = moat.get("conviction_score", 0) or 0
            tendencia = update_tracking(r["ticker"], drop_pct, vale_bool, conviction_int, tracking)
            r["tendencia"] = tendencia

            time.sleep(BATCH_DELAY)

        # Save tracking
        save_tracking(tracking)

        # Sugerencias — use active version if Perplexity key available
        print("\n  Generando sugerencias de nuevas empresas...")
        existing_names = [c["name"] for c in all_companies]
        if perplexity_key:
            suggestions = suggest_new_companies_active(client, perplexity_key, existing_names)
        else:
            suggestions = suggest_new_companies(client, existing_names)
        print(f"  {len(suggestions)} sugerencias generadas")

        # Analizar sugerencias con el mismo rigor que el universo
        print("\n  Analizando sugerencias con datos reales de mercado...")
        analyzed_suggestions = []
        for s in suggestions:
            ticker_yf = s.get("ticker_yahoo", "").strip()
            if not ticker_yf:
                continue
            sig = get_derating_signals(ticker_yf, ev_history)
            passes = passes_derating_threshold(sig, threshold)
            if "error" in sig or not passes:
                continue
            # Construir un objeto compatible con analyze_moat
            company_obj = {
                "name":    s.get("nombre", ""),
                "sector":  s.get("sector", ""),
                "geo":     s.get("geografia", ""),
            }
            moat = analyze_moat(client, company_obj, sig)
            if moat.get("vale_investigar") and moat.get("moat_strength") == "wide":
                s["_signals"] = sig
                s["_moat"]    = moat
                analyzed_suggestions.append(s)
                vale_str = moat.get("moat_type","")
                rev_str  = moat.get("reversibilidad","")
                drop     = sig.get("drop_from_high_pct","")
                print(f"    [SI] {s.get('nombre','')} — {drop}% caida | {vale_str} | {rev_str}")
            time.sleep(BATCH_DELAY)
        save_ev_history(ev_history)
        print(f"  {len(analyzed_suggestions)} sugerencias pasan el analisis completo")

        # Ranking de candidatas (universo)
        vale_list = [r for r in results if r.get("moat", {}).get("vale_investigar") is True]
        ranking = {}
        if vale_list:
            print("\n  Rankeando candidatas por fit con Estela Capital...")
            ranking = rank_candidates(client, vale_list)
            print(f"  Ranking generado para {len(ranking)} empresas")

        # Aplicar ranking a resultados
        for r in results:
            rank_info = ranking.get(r["name"], {})
            r["prioridad"]       = rank_info.get("prioridad", 99)
            r["destacada"]       = rank_info.get("destacada", False)
            r["razon_prioridad"] = rank_info.get("razon_prioridad", "")
    else:
        suggestions = []
        analyzed_suggestions = []
        for r in results:
            r["prioridad"]       = 99
            r["destacada"]       = False
            r["razon_prioridad"] = ""
            r["tendencia"]       = ""
        if args.skip_moat:
            print("\n[3/4] Análisis de moat omitido (--skip-moat)")

    # --- GLOBAL SCAN via FMP ---
    global_candidates_raw = []
    if fmp_key and not args.skip_moat and not args.skip_global:
        print("\n[GLOBAL SCAN] Buscando candidatas fuera del universo via FMP...")
        try:
            from fmp_screener import find_global_candidates
            import concurrent.futures as _cf
            with _cf.ThreadPoolExecutor(max_workers=1) as _ex:
                _fut = _ex.submit(find_global_candidates, fmp_key, threshold, 50)
                try:
                    global_candidates_raw = _fut.result(timeout=600)
                    print(f"  {len(global_candidates_raw)} candidatas globales encontradas")
                except _cf.TimeoutError:
                    print("  [WARN] Global scan timeout (>600s) — omitiendo esta semana")
                    global_candidates_raw = []
        except Exception as _fmp_err:
            print(f"  [WARN] FMP global scan falló: {_fmp_err}")
            global_candidates_raw = []

        if global_candidates_raw and client:
            print(f"  Analizando moats de candidatas globales con Claude...")
            for _gi, _gr in enumerate(global_candidates_raw, 1):
                print(f"    [{_gi:2d}/{len(global_candidates_raw)}] {_gr['name']}")
                _company_obj = {"name": _gr["name"], "sector": _gr["sector"], "geo": _gr["geo"]}
                _moat = analyze_moat(client, _company_obj, _gr["signals"])
                _gr["moat"] = _moat
                if "error" in _moat:
                    print(f"           [WARN] {_moat['error']}")
                else:
                    _v = "[SI]" if _moat.get("vale_investigar") else "[NO]"
                    _c = _moat.get("conviction_score", "?")
                    print(f"           {_v} | {_moat.get('moat_type','')} | {_moat.get('reversibilidad','')} | conviction: {_c}/10")
                _gr["tendencia"] = ""
                _gr["prioridad"] = 99
                _gr["destacada"] = False
                _gr["razon_prioridad"] = ""
                time.sleep(BATCH_DELAY)
        elif global_candidates_raw and not client:
            # skip_moat or no key — attach empty moat
            for _gr in global_candidates_raw:
                _gr["moat"] = {}
                _gr["tendencia"] = ""
                _gr["prioridad"] = 99
                _gr["destacada"] = False
                _gr["razon_prioridad"] = ""

    # Tag universe results with source, add global candidates to results
    for r in results:
        r.setdefault("source", "universo")
    for _gr in global_candidates_raw:
        _gr.setdefault("source", "fmp_global")
    results.extend(global_candidates_raw)

    # --- PASO 3: Ordenar resultados ---
    def sort_key(r):
        vale = r.get("moat", {}).get("vale_investigar")
        prio = r.get("prioridad", 99)
        drop = r.get("signals", {}).get("drop_from_high_pct", 0) or 0
        if vale is True:
            return (0, prio, -drop)
        elif vale is False:
            return (1, 99, -drop)
        else:
            return (2, 99, -drop)

    results.sort(key=sort_key)

    # --- PASO 4: Generar Excel ---
    print("\n[4/4] Generando Excel...")
    today = date.today().strftime("%Y-%m-%d")
    output_path = OUTPUT_DIR / f"screener_output_{today}.xlsx"
    # If file is locked (open in Excel), append a counter
    counter = 1
    while True:
        try:
            build_excel(results, suggestions, output_path)
            break
        except PermissionError:
            output_path = OUTPUT_DIR / f"screener_output_{today}_{counter}.xlsx"
            counter += 1

    # --- Web dashboard (GitHub Pages) ---
    try:
        docs_dir = BASE_DIR / "docs"
        docs_dir.mkdir(exist_ok=True)
        dashboard_html = generate_web_dashboard(results, proposed, today)
        (docs_dir / "index.html").write_text(dashboard_html, encoding="utf-8")
        print(f"\n[OK] Dashboard web generado en: docs/index.html")
    except Exception as _dash_err:
        print(f"\n[WARN] Dashboard web falló: {_dash_err}")

    # Resumen
    vale_investigar = [r for r in results if r.get("moat", {}).get("vale_investigar") is True]
    print(f"\n{'='*60}")
    print(f"  RESUMEN")
    print(f"{'='*60}")
    print(f"  Universo analizado:      {len(results)}")
    print(f"  Con derating >={threshold}%:    {len(candidates)}")
    print(f"  Vale investigar:         {len(vale_investigar)}")
    print(f"  Output:                  {output_path.name}")
    print(f"{'='*60}")

    # --- Proponer al universo (requiere confirmacion via confirm.py) ---
    proposed = []
    if not args.skip_moat:
        proposed = propose_to_universe(analyzed_suggestions, COMPANIES_JSON)
        if proposed:
            print(f"\n  [{len(proposed)} empresa(s) propuesta(s) pendientes de confirmacion]")
            for p in proposed:
                print(f"    ? {p['nombre']} ({p['ticker_yahoo']}) — {p['moat_type']} | {p.get('moat_strength','')}")

    # Enviar email
    if not args.skip_moat:
        print("\n[5/5] Enviando email...")
        send_email(output_path, results, threshold, proposed,
                   perplexity_key=perplexity_key, global_candidates=global_candidates_raw,
                   cloud=args.cloud)

    # Abrir el fichero automáticamente en Windows
    try:
        os.startfile(str(output_path))
    except Exception:
        pass


if __name__ == "__main__":
    main()
